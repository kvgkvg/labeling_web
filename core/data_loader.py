"""
Load and merge all XLSX files from data/{bbox,reasoning,answer}/ into memory.

Merge key: (dataset_name, row_position)
  - dataset_name  = raw filename stem (full, e.g. "GeminiFlashLite2-5_BLINK")
  - row_position  = 1-based row index in the raw file (header = row 0)
  - Filter file Row column = same 1-based position

CRITICAL: Excel caps sheet names at 31 chars.
  Match filter sheets via: full_stem[:31] == sheet_name
"""

import json
import logging
import os
import re
from glob import glob
from pathlib import Path
from typing import Any, Optional

import openpyxl

from core.bbox_parser import parse_bbox

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Dataset schema — describes per-dataset column layout
# ---------------------------------------------------------------------------
DATASET_SCHEMA: dict[str, dict] = {
    "GeminiFlashLite2-5_BLINK": {
        "question": "question",
        "image_path": "image_path",      # JSON list of filenames
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": ["A", "B", "C", "D"],
    },
    "GeminiFlashLite2-5_MMBench_DEV_EN_V11": {
        "question": "question",
        "image_path": None,              # no image_path col; use index value
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": ["A", "B", "C", "D"],
    },
    "GeminiFlashLite2-5_MME-RealWorld": {
        "question": "question",
        "image_path": "image_path",      # absolute path string from another machine
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": ["A", "B", "C", "D", "E"],
    },
    "GeminiFlashLite2-5_MME": {
        "question": "question",
        "image_path": "image_path",      # relative path string
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": [],                   # yes/no answers, no ABCD columns
    },
    "GeminiFlashLite2-5_MMStar": {
        "question": "question",
        "image_path": None,              # no image_path col; use index value
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": ["A", "B", "C", "D"],
    },
    "GeminiFlashLite2-5_SEEDBench_IMG": {
        "question": "question",
        "image_path": None,              # no image_path col; use index value
        "groundtruth": "answer",
        "prediction": "prediction",
        "choices": ["A", "B", "C", "D"],
    },
}

_DEFAULT_SCHEMA = {
    "question": "question",
    "image_path": "image_path",
    "groundtruth": "answer",
    "prediction": "prediction",
    "choices": ["A", "B", "C", "D"],
}

# Global samples store
SAMPLES: dict[str, dict] = {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_xlsx(path: str) -> dict[str, list[list]]:
    """Read all sheets from an xlsx file. Returns {sheet_name: [rows]}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        result[name] = rows
    wb.close()
    return result


def _rows_to_dicts(rows: list[list]) -> list[dict]:
    """Convert header + data rows into list of dicts."""
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        # Pad row if shorter than headers
        padded = list(row) + [None] * (len(headers) - len(row))
        result.append(dict(zip(headers, padded)))
    return result


def _parse_images(raw_value: Any, schema: dict, index_value: Any) -> list[str]:
    """
    Extract list of image basenames from raw_value per dataset schema rules.
    """
    if schema.get("image_path") is None:
        # No image_path column — use index column value as the image filename
        return [str(index_value)] if index_value is not None else []

    if raw_value is None:
        return []

    s = str(raw_value).strip()

    # Try JSON list first (BLINK sometimes has valid JSON)
    try:
        parsed = json.loads(s)
        if isinstance(parsed, list):
            return [os.path.basename(str(x)) for x in parsed if x]
    except (json.JSONDecodeError, ValueError):
        pass

    # Try ast.literal_eval (BLINK uses single-quote Python list repr)
    import ast as _ast
    try:
        parsed = _ast.literal_eval(s)
        if isinstance(parsed, list):
            return [os.path.basename(str(x)) for x in parsed if x]
    except (ValueError, SyntaxError):
        pass

    # Single path string (MME, MME-RealWorld)
    return [os.path.basename(s)] if s else []


def _val(v: Any) -> Optional[str]:
    """Convert cell value to string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s not in ("None", "nan", "") else None


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------

def load_all(data_dir: str = "data") -> dict[str, dict]:
    """
    Load and merge all XLSX files. Populates global SAMPLES dict.
    Returns the SAMPLES dict.
    """
    global SAMPLES
    raw_records: dict[tuple, dict] = {}   # (dataset_name, row_pos) → record

    # ------------------------------------------------------------------ #
    # Step 1: Load raw files for each type                                 #
    # ------------------------------------------------------------------ #
    for ftype in ("bbox", "reasoning", "answer"):
        type_dir = os.path.join(data_dir, ftype)
        filter_name = f"{ftype}_filter.xlsx"
        pattern = os.path.join(type_dir, "*.xlsx")

        raw_files = [
            f for f in glob(pattern)
            if os.path.basename(f) != filter_name
        ]

        for raw_path in sorted(raw_files):
            dataset_name = Path(raw_path).stem
            schema = DATASET_SCHEMA.get(dataset_name, _DEFAULT_SCHEMA)

            sheets = _read_xlsx(raw_path)
            sheet_rows = next(iter(sheets.values()))  # raw files have 1 sheet
            row_dicts = _rows_to_dicts(sheet_rows)

            for row_pos, row in enumerate(row_dicts, start=1):
                key = (dataset_name, row_pos)
                if key not in raw_records:
                    raw_records[key] = {
                        "dataset": dataset_name,
                        "row_idx": row_pos,
                        "id": f"{dataset_name}_{row_pos}",
                    }

                rec = raw_records[key]

                if ftype == "bbox":
                    # Only bbox raw carries the core sample data
                    q_col = schema.get("question", "question")
                    gt_col = schema.get("groundtruth", "answer")
                    pred_col = schema.get("prediction", "prediction")
                    img_col = schema.get("image_path")
                    choices_cols = schema.get("choices", ["A", "B", "C", "D"])

                    rec["question"] = _val(row.get(q_col))
                    rec["answer"] = _val(row.get(gt_col))
                    rec["prediction_raw"] = _val(row.get(pred_col))
                    rec["category"] = _val(row.get("category"))

                    # index column (display metadata, not merge key)
                    index_val = row.get("index")

                    # image_path
                    img_raw = row.get(img_col) if img_col else None
                    rec["images"] = _parse_images(img_raw, schema, index_val)

                    # MCQ choices
                    rec["choices"] = {
                        c: _val(row.get(c))
                        for c in choices_cols
                        if _val(row.get(c)) is not None
                    }

        # ------------------------------------------------------------------ #
        # Step 2: Load filter file and merge                                  #
        # ------------------------------------------------------------------ #
        filter_path = os.path.join(type_dir, filter_name)
        if not os.path.exists(filter_path):
            logger.warning(f"Filter file not found: {filter_path}")
            continue

        # Build reverse map: truncated_stem[:31] → full dataset_name
        stem_map: dict[str, str] = {}
        for ds in DATASET_SCHEMA:
            stem_map[ds[:31]] = ds

        filter_sheets = _read_xlsx(filter_path)

        for sheet_name, sheet_rows in filter_sheets.items():
            full_name = stem_map.get(sheet_name)
            if full_name is None:
                logger.warning(f"No matching dataset for filter sheet '{sheet_name}' in {filter_name}")
                continue

            filter_rows = _rows_to_dicts(sheet_rows)
            for frow in filter_rows:
                row_val = frow.get("Row")
                if row_val is None:
                    continue
                try:
                    row_pos = int(row_val)
                except (ValueError, TypeError):
                    continue

                key = (full_name, row_pos)
                if key not in raw_records:
                    raw_records[key] = {
                        "dataset": full_name,
                        "row_idx": row_pos,
                        "id": f"{full_name}_{row_pos}",
                    }

                rec = raw_records[key]
                status = _val(frow.get("Status"))

                if ftype == "bbox":
                    rec["bbox_status"] = status
                    rec["visual_reasoning"] = _val(frow.get("extracted_reasoning"))
                    rec["extracted_bbox"] = parse_bbox(frow.get("extracted_bbox"))
                    rec["bbox_answer"] = _val(frow.get("extracted_answer"))

                elif ftype == "reasoning":
                    rec["reasoning_status"] = status
                    rec["reasoning"] = _val(frow.get("extracted_reasoning"))
                    rec["reason_answer"] = _val(frow.get("extracted_answer"))
                    # "Steps found" column is ignored

                elif ftype == "answer":
                    rec["answer_status"] = status
                    rec["mcq_answer"] = _val(frow.get("extracted_answer"))
                    # "prediction (preview)" is discarded

    # ------------------------------------------------------------------ #
    # Step 3: Compute is_valid and build SAMPLES                          #
    # ------------------------------------------------------------------ #
    SAMPLES = {}
    for key, rec in raw_records.items():
        bbox_ok = rec.get("bbox_status") == "Valid"
        reason_ok = rec.get("reasoning_status") == "Valid"
        # answer_status has variants like "Valid (single_letter)", "Valid (boxed)", etc.
        answer_ok = str(rec.get("answer_status") or "").startswith("Valid")
        rec["is_valid"] = bbox_ok and reason_ok and answer_ok

        # Ensure all expected fields have defaults
        rec.setdefault("question", None)
        rec.setdefault("answer", None)
        rec.setdefault("images", [])
        rec.setdefault("choices", {})
        rec.setdefault("category", None)
        rec.setdefault("prediction_raw", None)
        rec.setdefault("visual_reasoning", None)
        rec.setdefault("extracted_bbox", [])
        rec.setdefault("bbox_answer", None)
        rec.setdefault("reasoning", None)
        rec.setdefault("reason_answer", None)
        rec.setdefault("mcq_answer", None)
        rec.setdefault("bbox_status", None)
        rec.setdefault("reasoning_status", None)
        rec.setdefault("answer_status", None)

        SAMPLES[rec["id"]] = rec

    valid_count = sum(1 for r in SAMPLES.values() if r["is_valid"])
    logger.info(f"Loaded {len(SAMPLES)} total samples, {valid_count} valid")
    return SAMPLES


def get_sample(sample_id: str) -> Optional[dict]:
    return SAMPLES.get(sample_id)


def list_samples() -> dict[str, dict]:
    return SAMPLES
