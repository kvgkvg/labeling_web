"""
Generate synthetic test data matching the real VQA dataset schema.

Creates:
  data/bbox/GeminiFlashLite2-5_{Dataset}.xlsx       (6 raw files)
  data/bbox/bbox_filter.xlsx                         (6-sheet filter)
  data/reasoning/GeminiFlashLite2-5_{Dataset}.xlsx  (6 raw files)
  data/reasoning/reasoning_filter.xlsx               (6-sheet filter)
  data/answer/GeminiFlashLite2-5_{Dataset}.xlsx     (6 raw files)
  data/answer/answer_filter.xlsx                     (6-sheet filter)
  ~/LMUData/{DatasetName}/*.jpg                      (placeholder images)

Run from project root:
  conda activate label_website
  python scripts/generate_sample_data.py
"""

import ast
import json
import os
import random
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl pillow")
    sys.exit(1)

random.seed(42)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROWS_PER_DATASET = 8   # synthetic rows per dataset
DATA_DIR = Path("data")
IMAGE_BASE_DIR = Path.home() / "LMUData"

QUESTIONS = [
    "Which image shows the same art style as the reference?",
    "What object is highlighted in the bounding box?",
    "Which of the following best describes the scene?",
    "What is the relationship between the two objects?",
    "Which answer correctly identifies the main subject?",
    "How many distinct regions are visible in the image?",
    "What color is the primary object in the image?",
    "Which option matches the visual pattern shown?",
]

CHOICE_TEXTS = {
    "A": ["The first image", "A red car", "Indoor scene", "Above", "Subject A"],
    "B": ["The second image", "A blue ball", "Outdoor scene", "Below", "Subject B"],
    "C": ["The third image", "A green tree", "Abstract scene", "Left of", "Subject C"],
    "D": ["The fourth image", "A yellow sign", "Aerial view", "Right of", "Subject D"],
    "E": ["None of the above", "Cannot determine", "All of the above", "Partially visible", "Other"],
}

CATEGORIES = ["Art_Style", "Spatial_Reasoning", "Object_Detection", "Scene_Understanding", "Pattern_Recognition"]

VISUAL_REASONINGS = [
    "Step 1: I analyze the image structure. Sub-step 1.1: The bounding box coordinates are extracted. Sub-step 1.2: The object inside the bbox is identified. Step 2: I compare with the reference. Conclusion: The answer is consistent with the visual evidence.",
    "Step 1: Sub-step 1.1: The question requires identifying spatial relationships. Sub-step 1.2: I locate both objects. Step 2: I determine their relative positions. Therefore: The spatial relationship is confirmed.",
    "Step 1: I examine the color and shape features. Sub-step 1.1: Primary color is identified. Sub-step 1.2: Shape contours are analyzed. Step 2: I match to the given options. Final answer: Option B matches best.",
]

REASONINGS = [
    "Step 1: Sub-step 1.1: I read the question carefully. Sub-step 1.2: I identify the key visual elements. Step 2: I evaluate each option systematically. Sub-step 2.1: Option A is ruled out. Sub-step 2.2: Option B matches. Conclusion: B is correct.",
    "Step 1: The question asks about a comparison. Sub-step 1.1: I analyze element 1. Sub-step 1.2: I analyze element 2. Step 2: I compare both elements. Therefore: The answer is C.",
    "Step 1: Sub-step 1.1: Context clues suggest option A. Sub-step 1.2: Visual confirmation obtained. Step 2: Final verification. Note: No ambiguity detected. Answer: A.",
]

PREDICTION_TEMPLATE = '{{"bbox": [{bbox}], "visual_reasoning": "{vr}", "reasoning": "{r}", "answer": "{ans}"}}'

# ---------------------------------------------------------------------------
# Dataset definitions
# ---------------------------------------------------------------------------

DATASETS = {
    "GeminiFlashLite2-5_BLINK": {
        "short": "BLINK",
        "choices": ["A", "B", "C", "D"],
        "image_path_mode": "json_list",   # ['img1.jpg', 'img2.jpg']
        "images_per_row": 2,
    },
    "GeminiFlashLite2-5_MMBench_DEV_EN_V11": {
        "short": "MMBench_DEV_EN_V11",
        "choices": ["A", "B", "C", "D"],
        "image_path_mode": "index",       # no image_path col; index = image stem
        "images_per_row": 1,
    },
    "GeminiFlashLite2-5_MME-RealWorld": {
        "short": "MME-RealWorld",
        "choices": ["A", "B", "C", "D", "E"],
        "image_path_mode": "single",      # single filename string
        "images_per_row": 1,
    },
    "GeminiFlashLite2-5_MME": {
        "short": "MME",
        "choices": [],                    # yes/no — no A/B/C/D columns
        "image_path_mode": "single",
        "images_per_row": 1,
    },
    "GeminiFlashLite2-5_MMStar": {
        "short": "MMStar",
        "choices": ["A", "B", "C", "D"],
        "image_path_mode": "index",
        "images_per_row": 1,
    },
    "GeminiFlashLite2-5_SEEDBench_IMG": {
        "short": "SEEDBench_IMG",
        "choices": ["A", "B", "C", "D"],
        "image_path_mode": "index",
        "images_per_row": 1,
    },
}


# ---------------------------------------------------------------------------
# Image helpers
# ---------------------------------------------------------------------------

PALETTE = [
    (180, 60, 60), (60, 120, 180), (60, 160, 80),
    (180, 140, 40), (120, 60, 180), (40, 160, 160),
]


def make_placeholder_image(path: Path, label: str, color: tuple):
    """Create a 480x360 placeholder JPEG with a colored background and label."""
    path.parent.mkdir(parents=True, exist_ok=True)
    img = Image.new("RGB", (480, 360), color=color)
    draw = ImageDraw.Draw(img)
    # Draw a lighter inner rectangle to simulate a bbox region
    draw.rectangle([60, 50, 380, 280], outline=(255, 255, 255), width=3)
    draw.text((10, 10), label, fill=(255, 255, 255))
    img.save(str(path), "JPEG")


def make_images_for_dataset(stem: str, short: str, mode: str, n: int, n_per_row: int) -> list:
    """
    Create placeholder images and return list of (image_path_value, [filenames]) per row.
    Returns list of (raw_image_path_cell_value, [basename, ...]) for each of n rows.
    """
    results = []
    for i in range(1, n + 1):
        filenames = []
        for j in range(1, n_per_row + 1):
            suffix = f"_{j}" if n_per_row > 1 else ""
            fname = f"{short}_{i}{suffix}.jpg"
            img_path = IMAGE_BASE_DIR / short / fname
            color = PALETTE[(i + j) % len(PALETTE)]
            make_placeholder_image(img_path, f"{short} row {i} img {j}", color)
            filenames.append(fname)

        if mode == "json_list":
            # BLINK: Python single-quote list repr
            cell = str(filenames)  # "['BLINK_1_1.jpg', 'BLINK_1_2.jpg']"
        elif mode == "single":
            cell = filenames[0]
        else:  # index — no image_path column, image resolved via index
            cell = None  # not written to sheet

        results.append((cell, filenames))
    return results


# ---------------------------------------------------------------------------
# Bbox helpers
# ---------------------------------------------------------------------------

def random_bbox_str() -> str:
    """Return a random bbox string in the '[x1, y1, x2, y2] {label}' format."""
    x1 = random.randint(50, 300)
    y1 = random.randint(50, 300)
    x2 = random.randint(x1 + 50, min(x1 + 400, 950))
    y2 = random.randint(y1 + 50, min(y1 + 400, 950))
    labels = ["object", "region", "target", ""]
    lbl = random.choice(labels)
    if lbl:
        return f'["[{x1}, {y1}, {x2}, {y2}] {{{lbl}}}"]'
    return f'["[{x1}, {y1}, {x2}, {y2}]"]'


def empty_bbox_str() -> str:
    return "[]"


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def build_raw_wb(stem: str, cfg: dict, image_rows: list, ftype: str) -> Workbook:
    """Build a raw XLSX workbook (single Sheet1) for one dataset + file type."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    choices = cfg["choices"]
    has_image_col = cfg["image_path_mode"] != "index"

    # Header
    headers = ["question", "index"]
    if has_image_col:
        headers.append("image_path")
    headers += ["answer", "category", "prediction"]
    headers += choices  # A, B, C, D (or A-E, or empty for MME)

    ws.append(headers)

    answers_pool = choices if choices else ["Yes", "No"]

    for i, (img_cell, filenames) in enumerate(image_rows, start=1):
        idx = filenames[0].replace(".jpg", "")  # use first filename stem as index
        gt = random.choice(answers_pool)
        q = QUESTIONS[i % len(QUESTIONS)]
        cat = CATEGORIES[i % len(CATEGORIES)]
        vr = VISUAL_REASONINGS[i % len(VISUAL_REASONINGS)]
        r = REASONINGS[i % len(REASONINGS)]
        bbox_raw = random_bbox_str() if random.random() > 0.2 else empty_bbox_str()
        pred = PREDICTION_TEMPLATE.format(
            bbox=f'"[100, 100, 400, 300]"',
            vr=vr[:60].replace('"', "'"),
            r=r[:60].replace('"', "'"),
            ans=gt,
        )

        row = [q, idx]
        if has_image_col:
            row.append(img_cell)
        row += [gt, cat, pred]
        for c in choices:
            row.append(random.choice(CHOICE_TEXTS.get(c, ["Option"])))

        ws.append(row)

    return wb


def build_filter_wb(datasets: dict, ftype: str, image_rows_map: dict) -> Workbook:
    """Build a filter XLSX workbook with one sheet per dataset (sheet name = stem[:31])."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    statuses = ["Valid", "Valid (single_letter)", "Invalid", "Không có trường bbox"]

    for stem, cfg in datasets.items():
        sheet_name = stem[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Row", "Status", "extracted_reasoning", "extracted_bbox", "extracted_answer"])

        image_rows = image_rows_map[stem]
        choices = cfg["choices"]
        answers_pool = choices if choices else ["Yes", "No"]

        for i in range(1, len(image_rows) + 1):
            # Make ~75% valid
            if random.random() < 0.75:
                status = "Valid" if ftype != "answer" else random.choice(["Valid", "Valid (single_letter)", "Valid (boxed)"])
            else:
                status = random.choice(["Invalid", "Không có trường bbox"])

            if ftype == "bbox":
                reasoning = VISUAL_REASONINGS[i % len(VISUAL_REASONINGS)]
                bbox = random_bbox_str() if random.random() > 0.2 else empty_bbox_str()
                answer = random.choice(answers_pool)
            elif ftype == "reasoning":
                reasoning = REASONINGS[i % len(REASONINGS)]
                bbox = empty_bbox_str()
                answer = random.choice(answers_pool)
            else:  # answer
                reasoning = ""
                bbox = empty_bbox_str()
                answer = random.choice(answers_pool)

            ws.append([i, status, reasoning, bbox, answer])

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Ensure we run from project root
    if not Path("data").exists():
        print("ERROR: Run from project root (label_website/), not scripts/")
        sys.exit(1)

    # Safety check — refuse to overwrite any existing data files
    existing = []
    for ftype in ("bbox", "reasoning", "answer"):
        type_dir = DATA_DIR / ftype
        for stem in DATASETS:
            p = type_dir / f"{stem}.xlsx"
            if p.exists():
                existing.append(str(p))
        for fname in (f"{ftype}_filter.xlsx",):
            p = type_dir / fname
            if p.exists():
                existing.append(str(p))

    if existing:
        print("ERROR: The following data files already exist and would be overwritten:")
        for f in existing:
            print(f"  {f}")
        print("\nMove or delete them first, then re-run this script.")
        print("WARNING: Do NOT delete real data — back it up first!")
        sys.exit(1)

    print("Generating synthetic test data...")

    # Build images and collect image_rows per dataset
    image_rows_map = {}
    for stem, cfg in DATASETS.items():
        print(f"  Creating images for {cfg['short']}...")
        image_rows_map[stem] = make_images_for_dataset(
            stem, cfg["short"], cfg["image_path_mode"],
            ROWS_PER_DATASET, cfg["images_per_row"]
        )

    # Write raw + filter files for each type
    for ftype in ("bbox", "reasoning", "answer"):
        type_dir = DATA_DIR / ftype
        type_dir.mkdir(parents=True, exist_ok=True)

        filter_sheets_wb = build_filter_wb(DATASETS, ftype, image_rows_map)
        filter_path = type_dir / f"{ftype}_filter.xlsx"
        filter_sheets_wb.save(str(filter_path))
        print(f"  Wrote {filter_path}")

        for stem, cfg in DATASETS.items():
            raw_wb = build_raw_wb(stem, cfg, image_rows_map[stem], ftype)
            raw_path = type_dir / f"{stem}.xlsx"
            raw_wb.save(str(raw_path))
            print(f"  Wrote {raw_path}")

    print("\nDone! Verify with:")
    print('  python -c "from core.data_loader import load_all; s=load_all(); print(len(s), \'total,\', sum(r[\'is_valid\'] for r in s.values()), \'valid\')"')


if __name__ == "__main__":
    main()
